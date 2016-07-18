def generate_SW0(date, cell):

        from openpyxl import load_workbook
        import shutil
        import os
        from kite_functions import day
        
        data_folder = 'C:\\FAO\\AquaCrop\\DATA\\'+str(cell)+'\\'
        
        steps=day(date)-day((1,1,1950))

        wb = load_workbook(filename=data_folder+str(cell)+'_SoilMoisture.xlsx', read_only=True)
        ws = wb['ACRU_Out']
        Percent_FC_A=float(ws.cell(column=4, row=2+steps).value)/ws.cell(column=6, row=2+steps).value
        Percent_FC_B=float(ws.cell(column=5, row=2+steps).value)/ws.cell(column=7, row=2+steps).value

        
        """
        found=False
        row=2
        while not found:
            if ws.cell(column=1, row=row).value==date[2]:
                print(ws.cell(column=1, row=row).value==date[2])
                if ws.cell(column=2, row=row).value==date[1]:
                    if ws.cell(column=3, row=row).value==date[0]:
                        Percent_FC_A=float(ws.cell(column=4, row=row).value)/ws.cell(column=6, row=row).value
                        Percent_FC_B=float(ws.cell(column=5, row=row).value)/ws.cell(column=7, row=row).value
                        found=True
                    else:
                        row+=1
                else:
                    row+=1
            else:
                row+=1
        """
        
	#date should take the form [{start sim}(m, d, y),{stop sim}(m,d,y),{start crop}(m,d,y,),{stop crop}(m,d,y)]
	data_folder = 'C:\\FAO\\AquaCrop\\DATA\\'+str(cell)+'\\'
        fc=open(data_folder+str(cell)+'_FieldCapacity.SW0', "r")        
        
	file = open(str(cell)+'_'+str(date[2])+'.SW0', "w")
	
	LINES=fc.readlines()

        SOIL_str=[line.split() for line in LINES[-4:]]
        
        SOIL=[[float(j) for j in i] for i in SOIL_str]
        
        SOIL[0][1]*=Percent_FC_A
        SOIL[1][1]*=Percent_FC_B
        SOIL[2][1]*=Percent_FC_B
        SOIL[3][1]*=Percent_FC_B
        
        for line in LINES[:12]:
	   file.write(line)
	
	for i in SOIL:
	   file.write('\t'+str(i[0])+'\t\t'+str(i[1])+'\t\t\t'+str(i[2])+'\n')
	
	file.close()
	shutil.move(os.getcwd()+'\\'+str(cell)+'_'+str(date[2])+'.SW0', 'C:\\FAO\\AquaCrop\\DATA\\'+str(cell)+'\\')